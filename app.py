import streamlit as st
import pandas as pd
import openpyxl
import io, tempfile, re
import traceback
from openpyxl.styles import PatternFill

st.set_page_config(page_title="더존 전표 변환기", page_icon="📊", layout="wide")
st.title("📊 더존 위하고 전표 변환기")


# ─────────────────────────────
# 숫자 변환
# ─────────────────────────────
def to_int(v):
    try:
        if v is None:
            return 0
        if isinstance(v, str):
            v = re.sub(r"[^0-9.-]", "", v)
        return int(float(v))
    except:
        return 0


def clean(v):
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except:
        pass
    return str(v).strip()


def parse_date(v):
    try:
        d = pd.to_datetime(v)
        return d.month, d.day
    except:
        return None, None


# ─────────────────────────────
# 거래유형 정규화
# ─────────────────────────────
def normalize_trade_type(ttype):
    ttype = str(ttype)

    if "매도" in ttype:
        return "SELL"
    elif "매수" in ttype:
        return "BUY"
    elif "예탁금" in ttype or "예탁금이용료" in ttype:
        return "INTEREST"
    elif "입고" in ttype or "공모주입고" in ttype:
        return "StockCredit"
    elif "입금" in ttype or "이체입금" in ttype:
        return "Credit"
    elif "출금" in ttype and (
        "은행이체" in ttype or
        "타사이체" in ttype or
        "이체출금" in ttype
    ):
        return "Debit"

    return None


# ─────────────────────────────
# row (엑셀 10컬럼 기준)
# ─────────────────────────────
def row(m, d, div, acct_code, acct_name, cp_code, cp_name, memo, dr, cr):
    return [m, d, div, acct_code, acct_name, cp_code, cp_name, memo, dr, cr]


# ─────────────────────────────
# 종목명 추출
# ─────────────────────────────
def extract_stock_name(name):
    name = str(name).replace(" ", "").strip()
    if "#" in name:
        return name.split("#")[-1]
    return name


# ─────────────────────────────
# 거래처 조회
# ─────────────────────────────
def load_broker_map(file):
    if file is None:
        return {}

    df = pd.read_excel(file)

    return {
        extract_stock_name(name): (   # 🔥 key = 종목명만
            str(code).strip(),       # 거래처코드
            str(name).strip()        # 원본 거래처명 유지
        )
        for code, name in zip(df.iloc[:, 0], df.iloc[:, 1])
    }


# ─────────────────────────────
# 매핑 조회 방식 변경
# ─────────────────────────────
def get_broker_info(stock, broker_map):

    key = extract_stock_name(stock)  # 🔥 동일한 기준으로 변환

    if key in broker_map:
        return broker_map[key]   # (code, name)
    else:
        return "", stock         # 미매핑


# ─────────────────────────────
# HANTOO 파서 (자동 컬럼 매칭)
# ─────────────────────────────
# ✅ HANTOO 파서: 위치 기반 컬럼 매핑
#    원본과 달라진 점:
#    1. data_start를 날짜 등장 행으로 동적 감지
#       → 헤더가 1줄이든 2줄이든 3줄이든 자동 대응
#    2. 헤더 전체(header_row ~ data_start-1)를 스캔해서
#       { 키워드: 컬럼위치 } col_map 생성
#       → 컬럼명 대신 인덱스로 값 접근 (ci_date, ci_type, ...)
#    3. 날짜 없는 보조 데이터줄을 buffer에 병합
#       → 거래단가/세금 등 2번째 줄 값도 누락 없이 흡수
# ─────────────────────────────
def parse_hantoo_sheet(df):

    # ① "거래일" 키워드가 있는 첫 번째 헤더 행 찾기
    header_row = None
    for i in range(min(15, len(df))):
        if any("거래일" in str(v or "") for v in df.iloc[i].astype(str)):
            header_row = i
            break

    if header_row is None:
        return []

    # 날짜가 처음 등장하는 행 = 데이터 시작점 (동적 감지)
    #       헤더 1줄 → data_start = header_row + 1
    #       헤더 2줄 → data_start = header_row + 2
    #       헤더 3줄 → data_start = header_row + 3  자동 처리
    data_start = None
    for i in range(header_row + 1, min(header_row + 10, len(df))):
        m, d = parse_date(df.iloc[i, 0])
        if m:
            data_start = i
            break

    if data_start is None:
        return []

    # 헤더 줄 전체 스캔 → { 키워드: 컬럼위치 } 매핑
    #       헤더가 몇 줄이든, 어느 줄에 키워드가 있든 위치로 기억
    col_map = {}
    for h_idx in range(header_row, data_start):
        for col_idx, val in enumerate(df.iloc[h_idx]):
            v = str(val).strip()
            if v and v != "nan":
                col_map[v] = col_idx

    # 키워드로 컬럼 위치(인덱스) 찾기
    #       🔥 수정: "정확히 일치"를 먼저 찾고, 없을 때만 "부분 일치"로 넘어감
    #       예) "구분" 키워드가 "대출구분" 같은 엉뚱한 컬럼에 먼저 걸리는 걸 방지
    def find_col_idx(keys):
        # 1순위: 컬럼명이 키워드와 정확히 같은 경우
        for k in keys:
            for col_name, idx in col_map.items():
                if k == col_name:
                    return idx
        # 2순위: 정확히 일치하는 게 없을 때만 부분 일치 허용
        for k in keys:
            for col_name, idx in col_map.items():
                if k in col_name:
                    return idx
        return None

    ci_date  = find_col_idx(["거래일", "거래일자", "일자", "날짜"])
    ci_type  = find_col_idx(["구분", "적요명", "내용", "거래종류", "거래명", "거래구분"])
    ci_stock = find_col_idx(["종목", "종목명"])
    ci_qty   = find_col_idx(["수량", "거래수량", "거래좌수"])
    ci_price = find_col_idx(["단가", "가격", "거래단가", "기준가"])
    ci_net   = find_col_idx(["금액", "거래금액", "입출금액", "정산금액", "거래대금"])
    ci_fee   = find_col_idx(["수수료/Fee", "수수료"])
    ci_tax   = find_col_idx(["tax", "세금", "제세금", "거래세/농특세", "거래세"])

    # 날짜 없는 보조 데이터줄을 바로 위 행(buffer)에 병합
    #       날짜 있는 행: 새 거래 시작 → buffer 교체
    #       날짜 없는 행: buffer의 빈 위치(None/""/0)만 채워넣기
    raw_rows = df.iloc[data_start:].reset_index(drop=True)
    merged_rows = []
    buffer = None

    for _, r in raw_rows.iterrows():
        m, d = parse_date(r.iloc[0])
        if m:
            if buffer is not None:
                merged_rows.append(buffer)
            buffer = list(r)
        else:
            if buffer is not None:
                for col_idx, v in enumerate(r):
                    if col_idx >= len(buffer):
                        continue
                    bv = buffer[col_idx]
                    is_empty = (bv is None or str(bv).strip() in ["", "nan", "0"])
                    if is_empty and str(v).strip() not in ["", "nan", "0"]:
                        buffer[col_idx] = v

    if buffer is not None:
        merged_rows.append(buffer)

    # 컬럼명 대신 위치 인덱스로 값 꺼내기
    def get(r, ci):
        return r[ci] if ci is not None and ci < len(r) else None

    trades = []
    for idx, r in enumerate(merged_rows):
        try:
            m, d = parse_date(get(r, ci_date))
            if not m:
                continue

            trade_type = clean(get(r, ci_type))
            stock      = clean(get(r, ci_stock)).strip()
            qty        = to_int(get(r, ci_qty))
            price      = to_int(get(r, ci_price))
            net        = to_int(get(r, ci_net))
            fee        = to_int(get(r, ci_fee))
            tax        = to_int(get(r, ci_tax))

            if not trade_type:
                continue

            trades.append({
                "month": m, "day": d, "type": trade_type,
                "stock": stock, "qty": qty, "price": price,
                "net": net, "fee": fee, "tax": tax
            })

        except Exception as e:
            st.error(f"❌ parse_hantoo_sheet 오류")
            st.write("행번호:", idx)
            st.write("원본 데이터:")
            st.write(r)
            st.write("에러:")
            st.write(str(e))
            st.code(traceback.format_exc())
            continue

    return trades


# ─────────────────────────────
# 전표 생성
# ─────────────────────────────
#                                      거래처코드,   예치금,  단기매매증권,  이자수익,        배당금수익,    적요명 뒤에 붙일 태그
def process_trades(trades, broker_map, broker_code, deposit, short_inv, interest_income, dividend_income, memo_suffix=""):
    rows = []

    for idx, t in enumerate(trades):
        try:
            m = t["month"]
            d = t["day"]
            type = t["type"]      # 구분, 적요명,내용,거래종류
            stock = t["stock"]    # 종목명
            stock_name = extract_stock_name(stock)    # 매핑용 거래처명
            ttype = normalize_trade_type(t["type"])
            qty = t["qty"]        # 수량
            price = t["price"]    # 단가
            net = t["net"]        # 거래금액
            fee = t["fee"]        # 거래수수료
            tax = t["tax"]        # 세금과공과

            if not ttype:
                continue

            # 🔥 매핑 적용
            cp_code, cp_name = get_broker_info(stock_name, broker_map)

            # 🔥 디버깅 (필요하면 주석 해제)
            # st.write("매핑확인:", stock, "→", cp_code, cp_name)

            # 매도
            if ttype == "SELL":
                memo = f"{stock_name}({qty}주*{price:,})매도{memo_suffix}"   # 🔥 태그 부착 + 단가 천단위 콤마: 예) ...매도#한투6716
                rows.append(row(m, d, "차변", deposit, "예치금", broker_code, "", memo, net, 0))
                rows.append(row(m, d, "대변", short_inv, "단기매매증권", cp_code, cp_name, memo, 0, qty * price))

            # 매수
            elif ttype == "BUY":
                cost = qty * price
                memo = f"{stock_name}({qty}주*{price:,})매수{memo_suffix}"   # 🔥 태그 부착 + 단가 천단위 콤마: 예) ...매수#한투6716
                rows.append(row(m, d, "차변", short_inv, "단기매매증권", cp_code, cp_name, memo, cost, 0))
                rows.append(row(m, d, "차변", 82800, "증권수수료", cp_code, cp_name, "매수수수료", fee, 0))
                rows.append(row(m, d, "대변", deposit, "예치금", broker_code, "", memo, 0, cost - fee))

            # 예탁금이용료
            elif ttype == "INTEREST":
                memo = "예탁금이용료"
                rows.append(row(m, d, "차변", deposit, "예치금", broker_code, "", memo, net, 0))
                rows.append(row(m, d, "대변", interest_income, "이자수익(금융)", broker_code, stock, memo, 0, net))

            # 공모주입고
            elif ttype == "StockCredit":
                cost = qty * price
                memo = f"{stock_name}({qty}주*{price:,})입고"
                rows.append(row(m, d, "차변", short_inv, "단기매매증권", cp_code, cp_name, memo, cost, 0))
                rows.append(row(m, d, "대변", 13100, "선급금", cp_code, cp_name, memo, 0, cost))

            # 이체입금
            elif ttype == "Credit":
                memo = f"{type}"
                rows.append(row(m, d, "차변", deposit, "예치금", broker_code, "", memo, net, 0))
                rows.append(row(m, d, "대변", deposit, "예치금", "", "미등록거래처", memo, 0, net))

            # 이체출금
            elif ttype == "Debit":
                memo = f"{type}"
                rows.append(row(m, d, "차변", deposit, "예치금", "", "미등록거래처", memo, 0, net))
                rows.append(row(m, d, "대변", deposit, "예치금", broker_code, "", memo, net, 0))

        except Exception as e:
            st.error(f"❌ process_trades 오류")
            st.write("거래번호:", idx)
            st.write("거래데이터:")
            st.write(t)
            st.write(str(e))
            st.code(traceback.format_exc())
            continue

    return rows


# ─────────────────────────────
# Excel 생성
# ─────────────────────────────
def create_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active

    header = [
        "1.월", "2.일", "3.구분",
        "4.계정과목코드", "5.계정과목명",
        "6.거래처코드", "7.거래처명",
        "8.적요명", "9.차변(출금)", "10.대변(입금)"
    ]

    ws.append(header)

    yellow_fill = PatternFill(
        start_color="FFF59D",
        end_color="FFF59D",
        fill_type="solid"
    )

    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).fill = yellow_fill

    for r in rows:
        ws.append(r)

    # 🔥 9.차변(출금), 10.대변(입금) 컬럼에 천단위 콤마 서식 적용 (예: 1000 → 1,000)
    for r_idx in range(2, ws.max_row + 1):
        ws.cell(row=r_idx, column=9).number_format = '#,##0'
        ws.cell(row=r_idx, column=10).number_format = '#,##0'

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ─────────────────────────────
# UI
# ─────────────────────────────
deposit = st.number_input("예치금", min_value=10000, max_value=99999, step=1, help="예: 12500")
short_inv = st.number_input("단기매매증권", min_value=10000, max_value=99999, step=1, help="예: 10700")
interest_income = st.number_input("이자수익(금융)", min_value=10000, max_value=99999, step=1, help="예: 42000")
dividend_income = st.number_input("배당금수익", min_value=10000, max_value=99999, step=1, help="예: 41800")

broker_file = st.file_uploader("거래처 매핑 엑셀 (이름 / 코드)", type=["xlsx"])

broker_code = st.number_input("증권사코드", min_value=90001, max_value=99999, step=1, help="예: 98001")

# 🔥 적요명 뒤에 붙일 태그 (사용자 입력)
#    예) memo_tag = "#한투6716" 입력 시 → "LG(6주*175800)매수#한투6716"
memo_tag = st.text_input("적요 뒤에 붙일 태그 (선택)", value="", help="예: #한투6716")

uploaded = st.file_uploader("엑셀 업로드")

if uploaded:
    if st.button("변환 실행"):

        broker_map = load_broker_map(broker_file) if broker_file else {}
        file_bytes = uploaded.getvalue()

        try:
            xl = pd.ExcelFile(io.BytesIO(file_bytes))
        except Exception as e:
            st.error("❌ 엑셀 읽기 실패")
            st.write(str(e))
            st.code(traceback.format_exc())
            st.stop()

        all_trades = []

        for sheet in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet, header=None)
            trades = parse_hantoo_sheet(df)
            all_trades.extend(trades)

        st.write("총 trades:", len(all_trades))
        st.write(all_trades)

        rows = process_trades(all_trades, broker_map, broker_code, deposit, short_inv, interest_income, dividend_income, memo_tag)

        preview_df = pd.DataFrame(
            rows,
            columns=["월", "일", "구분", "계정과목코드", "계정과목명", "거래처코드",
                     "거래처명", "적요명", "차변", "대변"]
        )

        st.subheader("📋 전표 미리보기")
        st.dataframe(preview_df, use_container_width=True, height=500)

        if not rows:
            st.error("❌ 변환 데이터 없음")
        else:
            out = create_excel(rows)
            st.success("완료")
            st.download_button("다운로드", data=out, file_name="result.xlsx")
